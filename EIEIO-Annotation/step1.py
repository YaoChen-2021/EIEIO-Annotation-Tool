import os
from pyopenms import MzMLFile, MSExperiment
from openpyxl import Workbook
def main(input_folder, split_folder):
    for file in os.listdir(input_folder):
        if file.endswith(".mzML"):
            file_path = os.path.join(input_folder, file)
            experiment = MSExperiment()
            MzMLFile().load(file_path, experiment)
            workbook = Workbook()
            extract_and_process(experiment, workbook)
            if "Sheet" in workbook.sheetnames:
                std = workbook["Sheet"]
                workbook.remove(std)
            split_workbook(workbook, os.path.join(split_folder, f"output_{file[:-5]}"))
def extract_and_process(experiment, workbook):
    j = 0
    for scan in experiment:
        if scan.getMSLevel() == 2:
            precursor = scan.getPrecursors()[0] if scan.getPrecursors() else None
            if precursor:
                precursor_mz = round(precursor.getMZ(), 4)
                precursor_rt = round(scan.getRT(), 4)
                sheet = workbook.create_sheet(title=f"mz_{j}_rt_{j}")
                sheet.append(["Precursor m/z", "Precursor RT", "MS2mz", "MS2i"])
                sheet.append([precursor_mz, precursor_rt, "", ""])
                for mz, i in zip(*scan.get_peaks()):
                    if i > 0:
                        mz = round(mz, 4)
                        i = round(i, 4)
                        sheet.append(["", "", mz, i])
                j += 1
def split_workbook(workbook, output_prefix):
    sheet_names = workbook.sheetnames
    batch_size = 2000
    current_workbook_index = 1
    current_sheet_count = 0
    output_workbook = Workbook()
    output_workbook.remove(output_workbook.active)
    for sheet_name in sheet_names:
        sheet = workbook[sheet_name]
        output_sheet = output_workbook.create_sheet(title=sheet_name)
        copy_sheet(sheet, output_sheet)
        current_sheet_count += 1
        if current_sheet_count == batch_size:
            output_file = f"{output_prefix}_{current_workbook_index}.xlsx"
            output_workbook.save(output_file)
            output_workbook = Workbook()
            output_workbook.remove(output_workbook.active)
            current_workbook_index += 1
            current_sheet_count = 0
    if current_sheet_count > 0:
        output_file = f"{output_prefix}_{current_workbook_index}.xlsx"
        output_workbook.save(output_file)
def copy_sheet(source_sheet, target_sheet):
    for row in source_sheet.iter_rows(values_only=True):
        target_sheet.append(row)

if __name__ == "__main__":
    input_folder = "input-mzml"
    split_folder = "output-step1"
    main(input_folder, split_folder)
